import os
import uuid
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment

from django.conf import settings
from django.core.paginator import Paginator
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.utils import timezone

from .models import OperationLog, MaintenanceLog, AlarmLog

User = get_user_model()


class LogService:
    """
    日志业务逻辑层
    """

    @staticmethod
    def _safe_int(value, default):
        """安全地将字符串转为 int，空值/空白/非法值返回 default"""
        if value is None:
            return default
        value = str(value).strip()
        if not value:
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _safe_bool(value):
        """安全地将字符串转为 bool，空值/空白返回 None"""
        if value is None:
            return None
        value = str(value).strip()
        if not value:
            return None
        return value.lower() in ('true', '1', 'yes')

    @staticmethod
    def _safe_str(value):
        """安全地获取字符串参数，空值/空白返回 None"""
        if value is None:
            return None
        value = str(value).strip()
        return value if value else None

    @staticmethod
    def get_operation_logs(request, user) -> dict:
        """
        获取人员操作日志列表

        Args:
            request: HTTP请求对象
            user: 当前用户对象

        Returns:
            dict: 包含分页数据的字典
        """
        page = LogService._safe_int(request.GET.get('page'), 1)
        limit = LogService._safe_int(request.GET.get('limit'), 10)
        start_time = LogService._safe_str(request.GET.get('start_time'))
        end_time = LogService._safe_str(request.GET.get('end_time'))
        action_type = LogService._safe_str(request.GET.get('action_type'))
        branch = LogService._safe_int(request.GET.get('branch'), None)
        maintenance_status = LogService._safe_str(request.GET.get('maintenance_status'))

        # 操作人筛选：兼容 user_id / operator / operator_name 三种参数名
        filter_user_id = LogService._safe_int(request.GET.get('user_id'), None)
        if filter_user_id is None:
            filter_user_id = LogService._safe_int(request.GET.get('operator'), None)
        if filter_user_id is None:
            operator_name = LogService._safe_str(request.GET.get('operator_name'))
            if operator_name:
                try:
                    filter_user_id = User.objects.get(username=operator_name).id
                except User.DoesNotExist:
                    filter_user_id = -1  # 不存在的用户 → 查不到任何结果
        filter_id = LogService._safe_int(request.GET.get('id'), None)

        queryset = OperationLog.objects.all()

        # 权限隔离：非管理员只能查看自己的操作日志
        if not user.is_superuser:
            queryset = queryset.filter(user_id=user.id)

        if start_time:
            queryset = queryset.filter(created_at__gte=start_time)
        if end_time:
            queryset = queryset.filter(created_at__lte=end_time)
        if action_type:
            queryset = queryset.filter(action_type=action_type)
        if filter_id is not None:
            queryset = queryset.filter(id=filter_id)
        if branch is not None:
            queryset = queryset.filter(branch=branch)
        if filter_user_id is not None:
            queryset = queryset.filter(user_id=filter_user_id)
        if maintenance_status:
            queryset = queryset.filter(maintenance_status=maintenance_status)

        queryset = queryset.order_by('-created_at')

        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)

        return {
            'count': paginator.count,
            'next': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'results': list(page_obj.object_list)
        }

    @staticmethod
    def get_maintenance_logs(request, user) -> dict:
        """
        获取故障处置日志列表

        Args:
            request: HTTP请求对象
            user: 当前用户对象

        Returns:
            dict: 包含分页数据的字典
        """
        page = LogService._safe_int(request.GET.get('page'), 1)
        limit = LogService._safe_int(request.GET.get('limit'), 10)
        start_time = LogService._safe_str(request.GET.get('start_time'))
        end_time = LogService._safe_str(request.GET.get('end_time'))
        fault_device = LogService._safe_str(request.GET.get('fault_device'))
        device_code = LogService._safe_str(request.GET.get('device_code'))
        filter_user_id = LogService._safe_int(request.GET.get('user_id'), None)

        queryset = MaintenanceLog.objects.all()

        # 权限隔离：非管理员只能查看自己的维修日志
        if not user.is_superuser:
            queryset = queryset.filter(user_id=user.id)

        if start_time:
            queryset = queryset.filter(created_at__gte=start_time)
        if end_time:
            queryset = queryset.filter(created_at__lte=end_time)
        if fault_device:
            queryset = queryset.filter(fault_device=fault_device)
        if device_code:
            # 通过关联告警的 branch 筛选
            alarm_ids = AlarmLog.objects.filter(branch=device_code).values_list('id', flat=True)
            queryset = queryset.filter(alarm_id__in=alarm_ids)
        if filter_user_id is not None:
            queryset = queryset.filter(user_id=filter_user_id)

        queryset = queryset.order_by('-created_at')

        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)

        return {
            'count': paginator.count,
            'next': page_obj.next_page_number() if page_obj.has_next() else None,
            'previous': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'results': list(page_obj.object_list)
        }

    @staticmethod
    def create_operation_log(request, user, validated_data) -> dict:
        """
        新增人员操作日志

        Args:
            request: HTTP请求对象
            user: 当前用户对象
            validated_data: 已验证的请求数据

        Returns:
            dict: 创建成功的记录数据
        """
        # 支持管理员替他人创建操作日志
        operator_user_id = validated_data.get('operator_user_id')
        if operator_user_id and user.is_staff:
            try:
                user = User.objects.get(id=operator_user_id)
            except User.DoesNotExist:
                pass  # 用户不存在则忽略，使用当前用户

        operation_log = OperationLog.objects.create(
            user_id=user.id,
            branch=validated_data['branch'],
            action_type=validated_data['action_type'],
            maintenance_status=validated_data['maintenance_status'],
            action_detail=validated_data.get('action_detail'),
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT'),
        )
        return {
            'id': operation_log.id,
            'user_id': operation_log.user_id,
            'branch': operation_log.branch,
            'action_type': operation_log.action_type,
            'maintenance_status': operation_log.maintenance_status,
            'action_detail': operation_log.action_detail,
            'created_at': operation_log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        }

    @staticmethod
    def create_maintenance_log(request, user, validated_data) -> dict:
        """
        新增故障处置日志

        Args:
            request: HTTP请求对象
            user: 当前用户对象
            validated_data: 已验证的请求数据

        Returns:
            dict: 创建成功的记录数据
        """
        maintenance_log = MaintenanceLog.objects.create(
            alarm_id=validated_data['alarm_id'],
            user_id=user.id,
            fault_date=timezone.now().date(),
            fault_time=timezone.now().time(),
            fault_device=validated_data.get('fault_device', ''),
            repair_detail=validated_data['repair_detail'],
            repair_images=validated_data.get('repair_images', []),
            remark=validated_data.get('remark', ''),
        )
        return {
            'id': maintenance_log.id,
            'alarm_id': maintenance_log.alarm_id,
            'user_id': maintenance_log.user_id,
            'fault_device': maintenance_log.fault_device,
            'repair_detail': maintenance_log.repair_detail,
            'repair_images': maintenance_log.repair_images,
            'remark': maintenance_log.remark,
            'created_at': maintenance_log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        }

    @staticmethod
    def upload_image(request, user) -> dict:
        """
        通用图片上传

        支持 jpg/png/webp 格式，最大 5MB。
        存储路径：media/{type}/{YYYY-MM-DD}/{uuid}.{ext}

        Args:
            request: HTTP请求对象（含 FILES）
            user: 当前用户对象

        Returns:
            dict: {"url": "/media/..."}

        Raises:
            ValueError: 文件类型不支持、文件过大或未上传文件
        """
        image_file = request.FILES.get('image')
        if not image_file:
            raise ValueError('未上传图片文件')

        # 验证文件类型
        ext = os.path.splitext(image_file.name)[1].lower()
        allowed_exts = {'.jpg', '.jpeg', '.png', '.webp'}
        if ext not in allowed_exts:
            raise ValueError(f'不支持的图片格式：{ext}，仅支持 jpg/png/webp')

        # 验证文件大小（最大 5MB）
        max_size = 5 * 1024 * 1024
        if image_file.size > max_size:
            raise ValueError('图片大小不能超过 5MB')

        # 构建存储路径
        upload_type = request.POST.get('type', 'maintenance')
        date_str = timezone.now().strftime('%Y-%m-%d')
        filename = f'{uuid.uuid4()}{ext}'
        relative_path = f'{upload_type}/{date_str}/{filename}'

        # 保存文件到磁盘
        save_path = os.path.join(settings.MEDIA_ROOT, relative_path)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        with open(save_path, 'wb+') as f:
            for chunk in image_file.chunks():
                f.write(chunk)

        return {'url': f'/media/{relative_path}'}

    @staticmethod
    def upload_to_oss(file_bytes: bytes, object_key: str) -> str:
        """
        上传文件到阿里云 OSS，返回 OSS object key。

        Args:
            file_bytes: 文件字节数据
            object_key: OSS 对象键，如 'alarm_images/1/2026-07-06/xxx.jpg'

        Returns:
            str: 上传成功的 OSS object key

        Raises:
            ValueError: OSS 配置缺失
            RuntimeError: 上传失败
        """
        oss_config = getattr(settings, 'OSS_CONFIG', None)
        if not oss_config or not oss_config.get('ACCESS_KEY_ID'):
            raise ValueError('OSS_CONFIG 未配置，无法上传')

        import oss2
        auth = oss2.Auth(oss_config['ACCESS_KEY_ID'], oss_config['ACCESS_KEY_SECRET'])
        bucket = oss2.Bucket(auth, oss_config['ENDPOINT'], oss_config['BUCKET_NAME'])

        result = bucket.put_object(object_key, BytesIO(file_bytes))
        if result.status != 200:
            raise RuntimeError(f'OSS 上传失败，状态码: {result.status}')

        return object_key

    @staticmethod
    def export_logs_to_excel(request, user, log_type, start_time, end_time):
        """
        导出日志为 Excel 文件

        Args:
            request: HTTP请求对象
            user: 当前用户对象
            log_type: 'operation' 或 'fault'
            start_time: 开始时间（可选）
            end_time: 结束时间（可选）

        Returns:
            HttpResponse: Content-Type 为 Excel 的响应对象
        """
        if log_type == 'operation':
            queryset = OperationLog.objects.all()
            if not user.is_superuser:
                queryset = queryset.filter(user_id=user.id)
            if start_time:
                queryset = queryset.filter(created_at__gte=start_time)
            if end_time:
                queryset = queryset.filter(created_at__lte=end_time)
            queryset = queryset.order_by('-created_at')

            # 生成 Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '人员操作日志'

            headers = ['操作日期', '操作时间', '操作人员', '操作支路', '操作类型', '运行状态', '详情']
            header_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, log in enumerate(queryset, 2):
                ws.cell(row=row_idx, column=1, value=log.created_at.strftime('%Y-%m-%d') if log.created_at else '')
                ws.cell(row=row_idx, column=2, value=log.created_at.strftime('%H:%M:%S') if log.created_at else '')
                ws.cell(row=row_idx, column=4, value=log.branch or '')
                ws.cell(row=row_idx, column=5, value=log.get_action_type_display())
                status_map = {'maintained': '已维护', 'maintaining': '维护中', 'unmaintained': '未维护'}
                ws.cell(row=row_idx, column=6, value=status_map.get(log.maintenance_status, log.maintenance_status))
                ws.cell(row=row_idx, column=7, value=str(log.action_detail) if log.action_detail else '')

                try:
                    operator = User.objects.get(id=log.user_id)
                    ws.cell(row=row_idx, column=3, value=operator.username)
                except User.DoesNotExist:
                    ws.cell(row=row_idx, column=3, value='未知用户')

            filename = f'operation_logs_{timezone.now().strftime("%Y%m%d")}.xlsx'

        elif log_type == 'fault':
            queryset = MaintenanceLog.objects.all()
            if not user.is_superuser:
                queryset = queryset.filter(user_id=user.id)
            if start_time:
                queryset = queryset.filter(created_at__gte=start_time)
            if end_time:
                queryset = queryset.filter(created_at__lte=end_time)
            queryset = queryset.order_by('-created_at')

            # 预加载 alarm 数据以减少查询
            alarm_ids = queryset.values_list('alarm_id', flat=True)
            alarm_map = {a.id: a for a in AlarmLog.objects.filter(id__in=alarm_ids)}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '故障处置日志'

            headers = ['故障日期', '故障时间', '故障设备', '故障支路', '维修人员', '备注', '图片']
            header_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, log in enumerate(queryset, 2):
                ws.cell(row=row_idx, column=1, value=log.created_at.strftime('%Y-%m-%d') if log.created_at else '')
                ws.cell(row=row_idx, column=2, value=log.created_at.strftime('%H:%M:%S') if log.created_at else '')
                ws.cell(row=row_idx, column=3, value=log.fault_device or '')
                ws.cell(row=row_idx, column=6, value=log.remark or '')
                ws.cell(row=row_idx, column=7, value=str(log.repair_images) if log.repair_images else '')

                alarm = alarm_map.get(log.alarm_id)
                ws.cell(row=row_idx, column=4, value=alarm.branch if alarm else '')

                try:
                    repairer = User.objects.get(id=log.user_id)
                    ws.cell(row=row_idx, column=5, value=repairer.username)
                except User.DoesNotExist:
                    ws.cell(row=row_idx, column=5, value='未知用户')

            filename = f'fault_logs_{timezone.now().strftime("%Y%m%d")}.xlsx'
        else:
            raise ValueError(f'不支持的日志类型：{log_type}')

        # 输出到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @staticmethod
    def get_log_statistics(request, user, date_str=None) -> dict:
        """
        获取日志统计概览

        返回今日和本月的告警、操作、处置统计数据。

        Args:
            request: HTTP请求对象
            user: 当前用户对象
            date_str: 统计日期 YYYY-MM-DD，默认今天

        Returns:
            dict: 包含 today 和 this_month 统计数据的字典
        """
        from datetime import datetime

        if date_str:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            target_date = timezone.now().date()

        # 本月第一天
        first_day_of_month = target_date.replace(day=1)

        # 今日统计
        today_operation_count = OperationLog.objects.filter(created_at__date=target_date).count()
        today_maintenance_count = MaintenanceLog.objects.filter(created_at__date=target_date).count()
        today_alarm_count = AlarmLog.objects.filter(timestamp__date=target_date).count()
        today_pending_alarm_count = AlarmLog.objects.filter(
            resolution_status='pending', timestamp__date=target_date
        ).count()

        # 本月统计
        month_operation_count = OperationLog.objects.filter(created_at__date__gte=first_day_of_month).count()
        month_maintenance_count = MaintenanceLog.objects.filter(created_at__date__gte=first_day_of_month).count()
        month_alarm_count = AlarmLog.objects.filter(timestamp__date__gte=first_day_of_month).count()
        month_pending_alarm_count = AlarmLog.objects.filter(
            resolution_status='pending', timestamp__date__gte=first_day_of_month
        ).count()

        return {
            'date': target_date.strftime('%Y-%m-%d'),
            'today': {
                'operation_count': today_operation_count,
                'maintenance_count': today_maintenance_count,
                'alarm_count': today_alarm_count,
                'pending_alarm_count': today_pending_alarm_count,
            },
            'this_month': {
                'operation_count': month_operation_count,
                'maintenance_count': month_maintenance_count,
                'alarm_count': month_alarm_count,
                'pending_alarm_count': month_pending_alarm_count,
            },
        }