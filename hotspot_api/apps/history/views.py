from __future__ import annotations

import json
import logging
import os
import time
from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.db import DatabaseError
from django.http import FileResponse
from django.core.exceptions import ObjectDoesNotExist
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from django.db.models import Q

from apps.common.response import Result
from apps.logs.models import MaintenanceLog
from .models import TemperatureRecord, AlarmRecord
from .serializers import (
    TemperatureRecordSerializer,
    TemperatureExportSerializer,
    AlarmRecordSerializer,
    AlarmExportSerializer,
)

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────
# 自定义分页 — 参数名匹配接口文档（page / size）
# ──────────────────────────────────────────────
class HistoryPagination(PageNumberPagination):
    page_query_param = 'page'
    page_size_query_param = 'size'
    page_size = 100
    max_page_size = 1000

    def get_paginated_response(self, data):
        return Result.success(data={
            'total': self.page.paginator.count,
            'page': self.page.number,
            'page_size': self.get_page_size(self.request),
            'list': data,
        })


# ──────────────────────────────────────────────
# 工具函数：解析 ISO 8601 时间参数
# ──────────────────────────────────────────────
def _parse_iso_datetime(value: str) -> datetime:
    """将 ISO 8601 字符串转为 datetime，支持带/不带时区尾缀"""
    if not value:
        raise ValueError('时间参数不能为空')
    # 兼容 "YYYY-MM-DDTHH:MM:SS" 和 "YYYY-MM-DD HH:MM:SS"
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(f'无法解析时间格式: {value}')


def _validate_time_range(start_time, end_time):
    """校验时间范围必填且有效，返回 (start_dt, end_dt) 或 Response"""
    if not start_time or not end_time:
        return (None, None), Result.error(code=400, msg='start_time 和 end_time 为必填参数')

    try:
        start_dt = _parse_iso_datetime(start_time)
        end_dt = _parse_iso_datetime(end_time)
    except ValueError as e:
        return (None, None), Result.error(code=400, msg=str(e))

    if start_dt >= end_dt:
        return (None, None), Result.error(code=400, msg='start_time 必须早于 end_time')

    return (start_dt, end_dt), None


def _check_admin(request):
    """校验管理员权限，非管理员返回 403 Response"""
    if not request.user.is_staff:
        return Result.error(code=403, msg='无权限执行此操作，仅管理员可导出')
    return None


# ╔══════════════════════════════════════════════════════════════╗
# ║               HistoryTemperatureViewSet                      ║
# ╚══════════════════════════════════════════════════════════════╝
class HistoryTemperatureViewSet(viewsets.GenericViewSet):
    """
    温度历史记录接口
    - GET  /api/history/temperature/         → 分页查询
    - POST /api/history/temperature/export/  → Excel 导出（仅管理员）
    """
    permission_classes = [IsAuthenticated]
    pagination_class = HistoryPagination
    serializer_class = TemperatureRecordSerializer

    # ── 异常判定阈值 ──
    ABNORMAL_TEMP_THRESHOLD = 80.0   # ℃
    ABNORMAL_AREA_THRESHOLD = 10.0   # %

    def _build_temperature_queryset(self, start_dt, end_dt, branch=None,
                                     max_temp=None, status=None,
                                     keyword=None):
        """构建温度查询过滤条件"""
        qs = TemperatureRecord.objects.filter(
            timestamp__gte=start_dt,
            timestamp__lte=end_dt,
        )
        if branch is not None:
            qs = qs.filter(branch=branch)
        if max_temp is not None:
            qs = qs.filter(max_temp__lte=max_temp)
        if status is not None:
            abnormal_q = Q(max_temp__gte=self.ABNORMAL_TEMP_THRESHOLD) | Q(area_ratio__gte=self.ABNORMAL_AREA_THRESHOLD)
            if status == 'abnormal':
                qs = qs.filter(abnormal_q)
            elif status == 'normal':
                qs = qs.exclude(abnormal_q)
        if keyword is not None:
            try:
                keyword_int = int(keyword)
                qs = qs.filter(Q(id=keyword_int) | Q(branch=keyword_int))
            except (TypeError, ValueError):
                pass  # 非数字 keyword 不匹配任何记录
        return qs.order_by('-timestamp')

    # ──────────────────────────────────────────
    # 3.1 历史温度查询
    # ──────────────────────────────────────────
    def list(self, request, *args, **kwargs):
        """
        GET /api/history/temperature/
        必填: start_time, end_time
        可选: branch, max_temp, status, keyword, page, size
        status: 正常 / 异常（筛选温度记录状态）
        keyword: 按记录 ID 或支路编号精确匹配（传入数字）
        """
        start_time = request.query_params.get('start_time')
        end_time = request.query_params.get('end_time')

        (start_dt, end_dt), err = _validate_time_range(start_time, end_time)
        if err:
            return err

        branch = request.query_params.get('branch')
        if branch is not None:
            try:
                branch = int(branch)
            except (TypeError, ValueError):
                return Result.error(code=400, msg='branch 参数格式错误')

        max_temp = request.query_params.get('max_temp')
        if max_temp is not None:
            try:
                max_temp = float(max_temp)
            except (TypeError, ValueError):
                return Result.error(code=400, msg='max_temp 参数格式错误')

        status = request.query_params.get('status')
        if status is not None:
            if status == '正常':
                status = 'normal'
            elif status == '异常':
                status = 'abnormal'
            else:
                return Result.error(code=400, msg='status 参数无效，可选: 正常 / 异常')

        keyword = request.query_params.get('keyword')

        queryset = self._build_temperature_queryset(
            start_dt, end_dt, branch, max_temp, status, keyword,
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Result.success(data={'total': queryset.count(), 'list': serializer.data})

    # ──────────────────────────────────────────
    # 3.2 温度统计汇总
    # ──────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """
        GET /history/temperature/summary/
        温度统计汇总，返回筛选条件下的全量统计（不翻页）。
        必填: start_time, end_time
        可选: branch, status
        """
        start_time = request.query_params.get('start_time')
        end_time = request.query_params.get('end_time')

        (start_dt, end_dt), err = _validate_time_range(start_time, end_time)
        if err:
            return err

        branch = request.query_params.get('branch')
        if branch is not None:
            try:
                branch = int(branch)
            except (TypeError, ValueError):
                return Result.error(code=400, msg='branch 参数格式错误')

        status = request.query_params.get('status')
        if status is not None:
            if status == '正常':
                status = 'normal'
            elif status == '异常':
                status = 'abnormal'
            else:
                return Result.error(code=400, msg='status 参数无效，可选: 正常 / 异常')

        queryset = self._build_temperature_queryset(start_dt, end_dt, branch, status=status)

        total = queryset.count()
        if total == 0:
            return Result.success(data={
                'total': 0,
                'avg_temp': 0.0,
                'max_temp': 0.0,
                'abnormal_count': 0,
            })

        from django.db.models import Avg, Max
        agg = queryset.aggregate(
            avg_temp=Avg('avg_temp'),
            max_temp=Max('max_temp'),
        )

        abnormal_q = Q(max_temp__gte=self.ABNORMAL_TEMP_THRESHOLD) | Q(area_ratio__gte=self.ABNORMAL_AREA_THRESHOLD)
        abnormal_count = queryset.filter(abnormal_q).count()

        return Result.success(data={
            'total': total,
            'avg_temp': round(float(agg['avg_temp'] or 0), 2),
            'max_temp': float(agg['max_temp'] or 0),
            'abnormal_count': abnormal_count,
        })

    # ──────────────────────────────────────────
    # 3.3 温度历史导出（仅管理员）
    # ──────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='export')
    def export(self, request):
        """
        POST /api/history/temperature/export/
        导出温度历史记录为 Excel 文件。
        异常温度（max_temp ≥ 80℃ 或 area_ratio ≥ 10%）单元格标红。
        """
        admin_err = _check_admin(request)
        if admin_err:
            return admin_err

        serializer = TemperatureExportSerializer(data=request.data)
        if not serializer.is_valid():
            return Result.error(code=400, msg='参数校验失败', data=serializer.errors)

        data = serializer.validated_data
        start_dt = data['start_time']
        end_dt = data['end_time']
        branch = data.get('branch')
        max_temp = data.get('max_temp')

        queryset = self._build_temperature_queryset(
            start_dt, end_dt, branch, max_temp,
        )

        return self._generate_temperature_excel(queryset, start_dt, end_dt)

    def _generate_temperature_excel(self, queryset, start_dt, end_dt):
        """使用 openpyxl 生成温度历史 Excel 文件流"""
        wb = Workbook()
        ws = wb.active
        ws.title = '温度历史记录'

        # ── 样式定义 ──
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
        red_fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
        red_font = Font(name='微软雅黑', color='FFFFFF', size=10, bold=True)
        normal_font = Font(name='微软雅黑', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # ── 表头 ──
        headers = ['支路编号', '采集时间', '最高温度(℃)', '平均温度(℃)', '热斑面积占比(%)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # ── 数据行 ──
        for row_idx, record in enumerate(queryset.iterator(chunk_size=500), 2):
            is_abnormal = (
                float(record.max_temp) >= self.ABNORMAL_TEMP_THRESHOLD
                or float(record.area_ratio) >= self.ABNORMAL_AREA_THRESHOLD
            )

            row_data = [
                record.branch,
                record.timestamp.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if record.timestamp else '',
                float(record.max_temp),
                float(record.avg_temp) if record.avg_temp is not None else '',
                float(record.area_ratio),
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = red_font if is_abnormal else normal_font
                cell.fill = red_fill if is_abnormal else PatternFill()
                cell.alignment = center_align
                cell.border = thin_border

        # ── 列宽自适应 ──
        col_widths = [10, 28, 14, 14, 16]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # ── 生成文件名 ──
        start_str = start_dt.strftime('%Y%m%d')
        end_str = end_dt.strftime('%Y%m%d')
        filename = f'温度记录_{start_str}_{end_str}.xlsx'

        return self._stream_excel(wb, filename)

    @staticmethod
    def _stream_excel(workbook, filename):
        """将 openpyxl Workbook 转为 FileResponse（文件下载响应）"""
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Length'] = buffer.getbuffer().nbytes
        return response


# ╔══════════════════════════════════════════════════════════════╗
# ║                 HistoryAlarmViewSet                          ║
# ╚══════════════════════════════════════════════════════════════╝
class HistoryAlarmViewSet(viewsets.GenericViewSet):
    """
    告警历史记录接口
    - GET  /api/history/alarm/              → 分页查询
    - GET  /api/history/alarm/{id}/presign/ → OSS 预签名 URL
    - POST /api/history/alarm/upload-image/ → 补偿上传故障图片到 OSS
    - POST /api/history/alarm/export/       → Excel 导出（仅管理员）
    """
    permission_classes = [IsAuthenticated]
    pagination_class = HistoryPagination
    serializer_class = AlarmRecordSerializer

    def _build_alarm_queryset(self, start_dt, end_dt, branch=None,
                               alarm_type=None, status=None,
                               keyword=None):
        """构建告警查询过滤条件"""
        qs = AlarmRecord.objects.filter(
            timestamp__gte=start_dt,
            timestamp__lte=end_dt,
        )
        if branch is not None:
            qs = qs.filter(branch=branch)
        if alarm_type is not None:
            qs = qs.filter(alarm_type=alarm_type)
        if status is not None:
            qs = qs.filter(status=status)
        if keyword is not None:
            try:
                keyword_int = int(keyword)
                qs = qs.filter(Q(id=keyword_int) | Q(branch=keyword_int))
            except (TypeError, ValueError):
                pass  # 非数字 keyword 不匹配任何记录
        return qs.order_by('-timestamp')

    def _get_maintenance_map(self, alarm_ids):
        """
        批量查询维修记录，返回 {alarm_id: MaintenanceLog} 映射。
        避免在 serializer 中逐条查询（N+1 问题）。
        使用 .only() 避免因数据库迁移滞后导致的缺失列错误。
        """
        if not alarm_ids:
            return {}
        maintenances = MaintenanceLog.objects.only(
            'alarm_id', 'id', 'user_id', 'repair_detail', 'repair_images',
        ).filter(alarm_id__in=alarm_ids)
        return {m.alarm_id: m for m in maintenances}

    # ──────────────────────────────────────────
    # 3.3 历史告警查询
    # ──────────────────────────────────────────
    def list(self, request, *args, **kwargs):
        """
        GET /api/history/alarm/
        必填: start_time, end_time
        可选: branch, alarm_type, status, keyword, page, size
        alarm_type: hot_spot / over_temp / offline
        keyword: 按记录 ID 或支路编号精确匹配（传入数字）
        """
        start_time = request.query_params.get('start_time')
        end_time = request.query_params.get('end_time')

        (start_dt, end_dt), err = _validate_time_range(start_time, end_time)
        if err:
            return err

        branch = request.query_params.get('branch')
        if branch is not None:
            try:
                branch = int(branch)
            except (TypeError, ValueError):
                return Result.error(code=400, msg='branch 参数格式错误')

        alarm_type = request.query_params.get('alarm_type')
        if alarm_type is not None:
            valid_types = {'hot_spot', 'over_temp', 'offline'}
            if alarm_type not in valid_types:
                return Result.error(code=400, msg='alarm_type 无效，可选: hot_spot/over_temp/offline')

        status = request.query_params.get('status')
        if status and status not in ('pending', 'resolved', 'recovering'):
            return Result.error(code=400, msg='status 参数无效，可选: pending/resolved/recovering')

        keyword = request.query_params.get('keyword')

        queryset = self._build_alarm_queryset(
            start_dt, end_dt, branch, alarm_type, status, keyword,
        )

        page = self.paginate_queryset(queryset)
        if page is not None:
            # 批量加载维修记录，注入 serializer context
            maintenance_map = self._get_maintenance_map([obj.id for obj in page])
            serializer = self.get_serializer(
                page, many=True,
                context={'maintenance_map': maintenance_map},
            )
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Result.success(data={'total': queryset.count(), 'list': serializer.data})

    # ──────────────────────────────────────────
    # 3.4 故障原图预签名 URL
    # ──────────────────────────────────────────
    @action(detail=True, methods=['get'], url_path='presign')
    def presign(self, request, pk=None):
        """
        GET /api/history/alarm/{id}/presign/
        根据告警记录 ID 生成 OSS 预签名 URL（有效期 300 秒）。
        """
        try:
            alarm = AlarmRecord.objects.get(pk=pk)
        except AlarmRecord.DoesNotExist:
            return Result.error(code=404, msg='告警记录不存在')

        object_key = alarm.image_path.lstrip('/') if alarm.image_path else ''

        url = self._generate_presigned_url(object_key)
        if url is None:
            return Result.error(code=500, msg='OSS 服务暂不可用，请稍后重试')

        return Result.success(data={'url': url})

    def _generate_presigned_url(self, object_key: str, expires: int = 300) -> str | None:
        """
        调用阿里云 OSS SDK 生成预签名 GET URL。
        返回 None 表示 OSS 配置缺失或 SDK 异常。
        """
        oss_config = getattr(settings, 'OSS_CONFIG', None)
        if not oss_config or not oss_config.get('ACCESS_KEY_ID'):
            logger.warning('OSS_CONFIG 未配置，无法生成预签名 URL')
            # 开发阶段回退：直接返回原始路径
            return object_key

        try:
            import oss2
            auth = oss2.Auth(oss_config['ACCESS_KEY_ID'], oss_config['ACCESS_KEY_SECRET'])
            bucket = oss2.Bucket(auth, oss_config['ENDPOINT'], oss_config['BUCKET_NAME'])
            url = bucket.sign_url('GET', object_key, expires)
            return url
        except ImportError:
            logger.error('oss2 SDK 未安装，无法生成预签名 URL')
            return None
        except Exception as e:
            logger.error(f'生成 OSS 预签名 URL 失败: {e}')
            return None

    # ──────────────────────────────────────────
    # 3.5 补偿上传 — 为指定告警记录上传故障图片到 OSS
    # ──────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='upload-image')
    def upload_alarm_image(self, request):
        """
        POST /api/history/alarm/upload-image/
        补偿上传：为指定告警记录上传故障图片到 OSS 并更新 image_path。

        请求体 (multipart/form-data):
            - alarm_id: int, 必填, 告警记录 ID
            - image: File, 必填, 图片文件 (jpg/png/webp, ≤5MB)
        """
        alarm_id = request.POST.get('alarm_id')
        image_file = request.FILES.get('image')

        # 参数校验
        if not alarm_id:
            return Result.error(code=400, msg='缺少必填参数: alarm_id')
        if not image_file:
            return Result.error(code=400, msg='未上传图片文件')
        try:
            alarm_id = int(alarm_id)
        except (ValueError, TypeError):
            return Result.error(code=400, msg='alarm_id 格式错误')

        try:
            alarm = AlarmRecord.objects.get(pk=alarm_id)
        except AlarmRecord.DoesNotExist:
            return Result.error(code=404, msg=f'告警记录不存在: id={alarm_id}')

        # 文件校验
        ext = os.path.splitext(image_file.name)[1].lower()
        allowed_exts = {'.jpg', '.jpeg', '.png', '.webp'}
        if ext not in allowed_exts:
            return Result.error(code=400, msg=f'不支持的图片格式：{ext}，仅支持 jpg/png/webp')
        if image_file.size > 5 * 1024 * 1024:
            return Result.error(code=400, msg='图片大小不能超过 5MB')

        # 读取文件字节
        image_bytes = image_file.read()

        # 生成 OSS object key
        from django.utils import timezone as tz
        now = tz.now()
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%Y%m%d_%H%M%S')
        branch = alarm.branch
        object_key = f'alarm_images/{branch}/{date_str}/{alarm_id}_{time_str}{ext}'

        # 上传到 OSS
        try:
            from apps.logs.services import LogService
            uploaded_key = LogService.upload_to_oss(image_bytes, object_key)
        except ValueError as e:
            return Result.error(code=500, msg=str(e))
        except RuntimeError as e:
            return Result.error(code=500, msg=str(e))

        # 更新告警记录的 image_path
        alarm.image_path = uploaded_key
        alarm.save(update_fields=['image_path'])

        logger.info(f'告警 {alarm_id} 的图片已补偿上传: {uploaded_key}')
        return Result.success(
            msg='图片上传成功',
            data={
                'alarm_id': alarm.id,
                'object_key': uploaded_key,
            },
        )

    # ──────────────────────────────────────────
    # 3.6 告警历史导出（仅管理员）
    # ──────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='export')
    def export_alarm(self, request):
        """
        POST /api/history/alarm/export/
        导出告警历史记录为 Excel 文件。
        """
        admin_err = _check_admin(request)
        if admin_err:
            return admin_err

        serializer = AlarmExportSerializer(data=request.data)
        if not serializer.is_valid():
            return Result.error(code=400, msg='参数校验失败', data=serializer.errors)

        data = serializer.validated_data
        start_dt = data['start_time']
        end_dt = data['end_time']
        branch = data.get('branch')
        alarm_type = data.get('alarm_type')
        status = data.get('status')

        queryset = self._build_alarm_queryset(
            start_dt, end_dt, branch, alarm_type, status,
        )

        return self._generate_alarm_excel(queryset, start_dt, end_dt)

    def _generate_alarm_excel(self, queryset, start_dt, end_dt):
        """使用 openpyxl 生成告警历史 Excel 文件流"""
        wb = Workbook()
        ws = wb.active
        ws.title = '告警历史记录'

        # ── 样式 ──
        header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
        cell_font = Font(name='微软雅黑', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # ── 表头 ──
        headers = [
            '支路编号', '告警类型', '触发时间',
            '温度(℃)', '面积占比(%)',
            '自动跳闸', '处置状态',
            '处置时间', '告警描述',
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # ── 数据行 ──
        status_display = {'pending': '未处理', 'resolved': '已处理', 'recovering': '恢复中'}

        for row_idx, record in enumerate(queryset.iterator(chunk_size=500), 2):
            row_data = [
                record.branch,
                record.get_alarm_type_display(),
                record.timestamp.strftime('%Y-%m-%d %H:%M:%S') if record.timestamp else '',
                float(record.temperature) if record.temperature is not None else '',
                float(record.area_ratio) if record.area_ratio is not None else '',
                '是' if record.auto_trip else '否',
                status_display.get(record.status, ''),
                record.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if record.resolved_at else '',
                record.description or '',
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = wrap_align if col_idx == len(headers) else center_align
                cell.border = thin_border

        # ── 列宽 ──
        col_widths = [10, 14, 22, 10, 12, 10, 10, 22, 40]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # ── 文件名 ──
        start_str = start_dt.strftime('%Y%m%d')
        end_str = end_dt.strftime('%Y%m%d')
        filename = f'报警记录_{start_str}_{end_str}.xlsx'

        return HistoryTemperatureViewSet._stream_excel(wb, filename)


# ╔══════════════════════════════════════════════════════════════╗
# ║             TemperatureStreamView  (SSE)                     ║
# ╚══════════════════════════════════════════════════════════════╝
class TemperatureStreamView(APIView):
    """
    实时温度 SSE（Server-Sent Events）推送
    GET /api/history/temperature/stream/

    每 2 秒推送 4 个支路的最新温度记录，
    每 30 秒发送心跳保活。
    """

    permission_classes = [IsAuthenticated]
    PUSH_INTERVAL = 2      # 数据推送间隔（秒）
    HEARTBEAT_INTERVAL = 30  # 心跳间隔（秒）

    @staticmethod
    def _fetch_latest(limit: int = 4):
        """
        从 temperature_records 按 timestamp 降序获取最新记录。
        每个 branch 各取一条，确保最多返回 4 条。
        """
        records = []
        seen_branches = set()
        qs = TemperatureRecord.objects.order_by('-timestamp')
        for record in qs.iterator(chunk_size=100):
            if record.branch not in seen_branches:
                records.append(record)
                seen_branches.add(record.branch)
            if len(records) >= limit:
                break
        return records

    @staticmethod
    def _format_record(r) -> dict:
        """将 ORM 对象转为 SSE 推送所需的字典格式"""
        return {
            'branch': r.branch,
            'max_temp': float(r.max_temp),
            'avg_temp': float(r.avg_temp) if r.avg_temp is not None else None,
            'area_ratio': float(r.area_ratio),
            'timestamp': r.timestamp.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] if r.timestamp else '',
        }

    def get(self, request, *args, **kwargs):
        def event_stream():
            last_heartbeat = time.monotonic()

            while True:
                try:
                    # ── 数据推送 ──
                    latest = self._fetch_latest()
                    if latest:
                        for record in latest:
                            payload = self._format_record(record)
                            line = json.dumps(payload, ensure_ascii=False)
                            yield f"event: status\ndata: {line}\n\n"
                    else:
                        # 表为空时推送空数组
                        yield f"event: status\ndata: []\n\n"

                    # ── 心跳 ──
                    now = time.monotonic()
                    if now - last_heartbeat >= self.HEARTBEAT_INTERVAL:
                        yield "event: heartbeat\ndata: \"ping\"\n\n"
                        last_heartbeat = now

                except DatabaseError as e:
                    logger.error(f'SSE 数据库查询失败: {e}')
                    yield f"event: error\ndata: {{\"error\": \"数据库查询失败\"}}\n\n"
                except Exception as e:
                    logger.error(f'SSE 未知异常: {type(e).__name__}: {e}')
                    yield f"event: error\ndata: {{\"error\": \"内部服务异常\"}}\n\n"

                time.sleep(self.PUSH_INTERVAL)

        response = StreamingHttpResponse(
            event_stream(),
            content_type='text/event-stream',
            status=200,
        )
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['X-Accel-Buffering'] = 'no'
        response['Access-Control-Allow-Origin'] = '*'
        response['Connection'] = 'keep-alive'
        return response
